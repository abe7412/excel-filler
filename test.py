import streamlit as st
import openai import OpenAI
import json
import pymupdf
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from io import BytesIO

# Function to safely write to merged cells
def safe_write(ws, cell, value):
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if ws[cell].coordinate in merged_range:
            cell = ws.cell(row=min_row, column=min_col).coordinate
            break
    ws[cell] = value

# Function to fill the Excel form
def fill_excel_form_custom(template_path, data, po_number, po_expiry):
    wb = load_workbook(template_path)
    ws = wb.active

    safe_write(ws, "E17", data["Name in English"])       # English name
    safe_write(ws, "S16", data["Nationality"])           # Nationality
    safe_write(ws, "AA16", data["Occupation in Arabic"])        # Arabic occupation
    safe_write(ws, "AD17", data["Occupation in English"])            # English occupation
    safe_write(ws, "A19", data["Iqama Number"])                 # Iqama number
    safe_write(ws, "G41", po_number)                     # PO Number
    safe_write(ws, "AE41", po_expiry)                    # PO Expiry
    safe_write(ws, "A16", data["Translated Name in Arabic"])
    safe_write(ws, "V17", data["Nationality"])           # Arabic name (reversed)

    return wb

# OpenAI Key
    apiKeys = st.secrets["API_Keys"]
    openAiKey = apiKeys["openAI"]
    client = OpenAI(api_key=openAiKey)


# Streamlit UI
st.title("Fill Excel Form with PDF Data")

# Input fields for the user
uploaded_pdf = st.file_uploader("Upload PDF", type="pdf")
uploaded_excel = st.file_uploader('Upload Excel', type = ['.xlsx','.xls'])
po_number = st.text_input("PO Number")
po_expiry = st.text_input("PO Expiry Date")

if uploaded_pdf and uploaded_excel:
    # Read the PDF and extract text
    stream = BytesIO(uploaded_pdf.read())
    doc = pymupdf.open(stream = stream)
    raw_text = ''
    for i in range(doc.page_count):
        page = doc.load_page(i)
        raw_text += page.get_text()
    


    prompt = '''given the following text, provide the following information in a dictionary format:
    1) Iqama Number
    2) Nationality
    3) Occupation in English
    4) Occupation in Arabic. If not provided, translate the English occupation
    5) Name in English
    6) Translated Name in Arabic
    7) Passport Expiry Date
    8) Iqama Expiry date'''

    prompt = prompt + raw_text

    # Generate response from OpenAI
    messages = [{"role": "user", "content": prompt}]
    response = client.ChatCompletion.create(
        model="gpt-4",
        messages=messages,
        max_tokens=500,
        temperature=0.8
    )

    summary = response.choices[0].message['content']
    summary = json.loads(summary)

    # Excel Template path
    excel_template = uploaded_excel.read()  # Adjust to your template path
    
    # Fill Excel form
    wb = fill_excel_form_custom(excel_template, summary, po_number, po_expiry)

    # Save to BytesIO for download
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Provide download link
    st.download_button(
        label="Download Filled Application",
        data=output,
        file_name="filled_application.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )